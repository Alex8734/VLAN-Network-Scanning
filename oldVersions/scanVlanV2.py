import csv
import requests
import subprocess
from bs4 import BeautifulSoup
import ssl
import socket
import warnings
import concurrent.futures
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from urllib3.exceptions import InsecureRequestWarning
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

header = ['IP Adresse', 'Port 80', 'Port 443', 'Zertifikat vorhanden', 'System', 'ICMP ECHO', 'HTTP redirect auf HTTPS']

# Funktion zur Überprüfung eines Ports
def check_port(ip, port, protocol):
    if protocol == 'http':
        url = f'http://{ip}:{port}'
    elif protocol == 'https':
        url = f'https://{ip}:{port}'
    else:
        return 'Unbekanntes Protokoll'
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'})


    try:
        
        response = session.get(url, timeout=2, allow_redirects=True, verify=False)
        soup = BeautifulSoup(response.content, 'html.parser')  
        title = soup.find('title').text if soup.find('title') else '-'  
            
        cert_present = '-'
        if protocol == 'https':
            try:
                # Versuch, eine SSL-Verbindung aufzubauen und das Zertifikat zu erhalten
                context = ssl.create_default_context()
                with socket.create_connection((ip, port)) as sock:
                    with context.wrap_socket(sock, server_hostname=ip) as ssock:
                        cert = ssock.getpeercert()
                        cert_present = 'Ja' if cert else 'Nein'
            except Exception as e:
                cert_present = 'Nein'


        if response.status_code == 301 and protocol == 'http':
            return ('Nein', 'Ja', title,cert_present)
        elif response.status_code in [200, 401]:
            return ('Ja', 'Ja', title,cert_present)
        else:
            return ('Nein', 'Nein', "-",cert_present)
    except requests.exceptions.RequestException:
        return ('Nein', 'Nein', "-","Nein")
    finally:
        session.close()


def check_ip(ip):
    ping_process = subprocess.run(['ping', '-n', '2', '-w', '1000', ip], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    reachable = 'Ja' if ping_process.returncode == 0 else 'Nein'
    
    port80, redirect, system80, _ = check_port(ip, 80, 'http')
    port443, _, system443, certAvailable = check_port(ip, 443, 'https')

    system = system80 if system80 != '-' else system443
    system = system.strip()



    if redirect == 'Ja':
        port80 = 'Nein'

    print(f'[Scanning] IP: {ip} --> done')
    return [ip, port80, port443, certAvailable, system, reachable, redirect]


def coloring(ws,results):
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    # Gelbe Füllung für keine offenen Ports 80 und 443
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    
    # Bedingte Formatierung für gelbe Markierung
    port_80_column = 'B'  # Angenommen, Port 80 ist in Spalte B
    port_443_column = 'C'  # Angenommen, Port 443 ist in Spalte C
    ws.conditional_formatting.add(f'A2:G{len(results) + 1}',
                                FormulaRule(formula=[f'AND(ISERROR(SEARCH("Ja",{port_80_column}2)), ISERROR(SEARCH("Ja",{port_443_column}2)))'], fill=yellow_fill))

    # Bedingte Formatierung für rote Markierung
    icmp_echo_column = 'F'  # Angenommen, die ICMP Echo-Antwort ist in Spalte F
    ws.conditional_formatting.add(f'{icmp_echo_column}2:{icmp_echo_column}{len(results) + 1}',
                                FormulaRule(formula=[f'ISERROR(SEARCH("Ja",{icmp_echo_column}2))'], fill=red_fill))


def create_table(ws,results,VLan):
    
    # Nachdem alle Daten hinzugefügt wurden
    table_ref = f"A1:G{len(results) + 1}"  # A1:G ist der Bereich der Tabelle, len(results) + 1 für die Header-Zeile
    table = Table(displayName=f"{VLan}ErgebnissTable", ref=table_ref)
    # Tabelle-Stil hinzufügen (Sie können den Stil anpassen)
    table_style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = table_style
    ws.add_table(table)

    # Zellenbreite automatisch anpassen
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

# Hauptprogramm
def main():
    VLan = '116'
    warnings.simplefilter('ignore', InsecureRequestWarning)
    
    try:
        wb = load_workbook('Ergebnisse.xlsx')
        ws_title = f'vlan {VLan}'
        if ws_title in wb.sheetnames:
            del wb[ws_title]
        ws = wb.create_sheet(title=ws_title)
        ws.append(header)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'vlan {VLan}'
        ws.append(header)
    
    ips = [f'192.168.{VLan}.{i}' for i in range(1, 256)]

    #multiprocessing
    with concurrent.futures.ThreadPoolExecutor() as executor:
        results = list(executor.map(check_ip, ips))
    
    for result in results:
        ws.append(result)
        print(f'[Saving] IP: {result[0]} --> done')
    
    create_table(ws,results,VLan)

    #coloring(ws,results)

    try:
        wb.save("Ergebnisse.xlsx")
    except PermissionError:
        print("Die Datei ist geöffnet. Bitte schließen Sie die Datei und versuchen Sie es erneut.")
        input("Drücken Sie eine beliebige Taste um es zu Wiederholen...")
        wb.save("Ergebnisse.xlsx")
    
if __name__ == "__main__":
    main()