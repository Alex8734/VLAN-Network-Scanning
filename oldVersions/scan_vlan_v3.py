import csv
import requests
import subprocess
from bs4 import BeautifulSoup
import ssl
import socket
import warnings
import concurrent.futures
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from urllib3.exceptions import InsecureRequestWarning
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from pysnmp.hlapi.asyncio.slim import Slim
from pysnmp.hlapi import SnmpEngine, CommunityData, UdpTransportTarget, ContextData, getCmd, ObjectType, ObjectIdentity
from openpyxl.formatting.rule import FormulaRule
import asyncio
import argparse


header = ['IP Adresse ', 'Port 80 ', 'Port 443 ', 'Zertifikat vorhanden ', 'System ', 'ICMP ECHO ', 'HTTP redirect auf HTTPS ','SNMP ', 'SNMP - sysDescr']


async def check_snmp(ip):
    with Slim(1) as slim:
        errorIndication, errorStatus, errorIndex, varBinds = await slim.get(
            "public",
            ip,
            161,
            ObjectType(ObjectIdentity("SNMPv2-MIB", "sysDescr", 0)),
            timeout=1,
            retries=0,
        )

        if errorIndication:
            return ""
        
        elif errorStatus:
            if errorStatus.prettyPrint() == 'noSuchName':
                return "Unknown"
            
            return "{} at {}".format(
                    errorStatus.prettyPrint(),
                    errorIndex and varBinds[int(errorIndex) - 1][0] or "?",
                )
            
        else:
            for varBind in varBinds:
                return varBind[1].prettyPrint()


def check_snmp_sync(ip):
    return asyncio.run(check_snmp(ip))


# Funktion zur Überprüfung eines Ports
def check_port(ip, port, protocol):
    if protocol == 'http':
        url = f'http://{ip}:{port}'
    elif protocol == 'https':
        url = f'https://{ip}:{port}'
    else:
        return 'Unbekanntes Protokoll'
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'})

    try:
        
        response = session.get(url, timeout=2, allow_redirects=True, verify=False)
        soup = BeautifulSoup(response.content, 'html.parser')  
        title = soup.find('title').text if soup.find('title') else '-'  ## nicht immer correct weil Titel im JS generiert wird.  HTML --> ("<!-- title is based on deviceName result of getPuDetails service -->\r\n <title></title>")
            
        cert_present = '-'
        if protocol == 'https':
            try:
                # Versuch, eine SSL-Verbindung aufzubauen und das Zertifikat zu erhalten
                context = ssl.create_default_context()
                with socket.create_connection((ip, port)) as sock:
                    with context.wrap_socket(sock, server_hostname=ip) as ssock:
                        cert = ssock.getpeercert()
                        cert_present = 'Ja' if cert else 'Nein'
            except Exception as e:
                cert_present = 'Nein'


        if response.status_code == 301 and protocol == 'http':
            return ('Nein', 'Ja', title,cert_present)
        
        return (response.status_code, 'Nein', title,cert_present)


    except requests.exceptions.RequestException:
        
        return ('-', 'Nein', "-","Nein")
    finally:
        session.close()


def check_ip(ip):
    ping_process = subprocess.run(['ping', '-n', '2', '-w', '1000', ip], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    reachable = 'Ja' if ping_process.returncode == 0 else 'Nein'
    
    port80 = port443 = snmp_devDesc = redirect = system = "-"
    certAvailable = snmp = "Nein"
    
    if reachable == 'Ja':
            
        port80, redirect, system80, _ = check_port(ip, 80, 'http')
        port443, _, system443, certAvailable = check_port(ip, 443, 'https')

        system = system80 if system80 != '-' else system443
        system = system.strip()
        snmp_devDesc = check_snmp_sync(ip)
        snmp = "Nein" if "" == snmp_devDesc else "Ja"

        if redirect == 'Ja':
            port80 = 'Nein'

    print(f'[Scanning] IP: {ip} --> done')

    return [ip, port80, port443, certAvailable, system, reachable, redirect,snmp, snmp_devDesc]


def coloring(ws,results):
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    # Gelbe Füllung für keine offenen Ports 80 und 443
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    
    # Bedingte Formatierung für gelbe Markierung
    port_80_column = 'B'  # Angenommen, Port 80 ist in Spalte B
    port_443_column = 'C'  # Angenommen, Port 443 ist in Spalte C
    ws.conditional_formatting.add(f'A2:G{len(results) + 1}',
                                FormulaRule(formula=[f'AND(ISERROR(SEARCH("Ja",{port_80_column}2)), ISERROR(SEARCH("Ja",{port_443_column}2)))'], fill=yellow_fill))

    # Bedingte Formatierung für rote Markierung
    icmp_echo_column = 'F'  # Angenommen, die ICMP Echo-Antwort ist in Spalte F
    ws.conditional_formatting.add(f'{icmp_echo_column}2:{icmp_echo_column}{len(results) + 1}',
                                FormulaRule(formula=[f'ISERROR(SEARCH("Ja",{icmp_echo_column}2))'], fill=red_fill))


def index_to_column(index):
    """Konvertiert einen 0-basierten Index in einen Spaltenbuchstaben."""
    column = ''
    while index >= 0:
        column = chr(index % 26 + 65) + column
        index = index // 26 - 1
    return column


def create_table(ws,results,VLan):
    header_length = len(header)  # Ersetze `header` mit deiner tatsächlichen Header-Variable
    last_column_letter = index_to_column(header_length - 1)
    
    # Nachdem alle Daten hinzugefügt wurden
    table_ref = f"A1:{last_column_letter}{len(results) + 1}"  # A1:len(header) ist der Bereich der Tabelle, len(results) + 1 für die Header-Zeile
    table = Table(displayName=f"_{VLan}ErgebnissTable", ref=table_ref)

    table.column_names.append([f"_{VLan}{header[i]}Column" for i in range(header_length)])
    # Tabelle-Stil hinzufügen (Sie können den Stil anpassen)
    table_style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = table_style
    ws.add_table(table)

    # Zellenbreite automatisch anpassen
    widths = [17,10.5,10.5,21,49,13,25,8,105]
    for i,col in enumerate(ws.columns):
        column = col[0].column  # Get the column name
        
        ws.column_dimensions[get_column_letter(column)].width = widths[i]


async def scan_vlan(VLan, startIdx, endIdx, wb, output_file):
    
    #excel stuff
    ws_title = f'vlan {VLan}'
    if ws_title in wb.sheetnames:
        del wb[ws_title]
    ws = wb.create_sheet(title=ws_title)
    ws.append(header)


    ips = [f'192.168.{VLan}.{i}' for i in range(int(startIdx), int(endIdx)+1)]

    #multiprocessing
    with concurrent.futures.ThreadPoolExecutor() as executor:
        results = list(executor.map(check_ip, ips))
    

    for result in results:
        ws.append(result)
        print(f'[Saving] IP: {result[0]} --> done')
    

    create_table(ws,results,VLan)
    try:
        wb.save(output_file)
    except PermissionError:
        pass
    #coloring(ws,results) # ist hässlich (funktioniert nicht wie geplant)


async def main():
    parser = argparse.ArgumentParser(description='Description of your program')
    parser.add_argument('-l', '--vlans',nargs="+", help='VLans to scan', required=False)
    parser.add_argument('-s', '--start', help='Start index', required=False)
    parser.add_argument('-e', '--end', help='End index', required=False)
    parser.add_argument('-f', '--file', help='Output file', required=False)
    args = parser.parse_args()
    
    file = 'Ergebnisse.xlsx'
    VLans = ['116','117']
    VLans = args.vlans if  args.vlans else VLans
    startIdx = args.start if args.start else 1
    endIdx = args.end if args.end else 255
    file = args.file if args.file else file

    if not file.endswith('.xlsx'):
        file += '.xlsx'
        
    warnings.simplefilter('ignore', InsecureRequestWarning)
    
    # excel stuff
    try:
        wb = load_workbook(file)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        
    for VLan in VLans:
        await scan_vlan(VLan, startIdx, endIdx, wb, file)

    while True:
        try:
            wb.save(file)
            break  # Wenn die Datei erfolgreich gespeichert wurde, beenden Sie die Schleife
        except PermissionError:
            print("Die Datei ist geöffnet. Bitte schließen Sie die Datei und versuchen Sie es erneut.")
            input("Drücken Sie eine beliebige Taste um es zu wiederholen...")
        except Exception as e:
            print(f"Ein unerwarteter Fehler ist aufgetreten: {e}")
            break  # Beenden Sie die Schleife bei einem unerwarteten Fehler
            
    subprocess.run(['start', file], shell=True)
    wb.close()
            
    
if __name__ == "__main__":
    asyncio.run(main())